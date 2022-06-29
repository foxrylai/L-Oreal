import os
import site
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import BORDER_DOUBLE
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

path = input('please input log files path \n')
path_excel = r'Y:\WorkFiles\Document\003_TLOC Group ID\check_files'
site_list = r'E:\Python\L-Oreal\TLOC Group ID\site_list.txt'
batch_list = r'E:\Python\L-Oreal\TLOC Group ID\batch_list.txt'
site_id_dict = {}
site_batch_dict = {}

files = os.listdir(path)
router_list = [i.split('.')[0].split('-')[0] for i in files]
router_list = list(set(router_list))

headlines = ['System IP','Site ID','STATE','Source TLOC Color','Remote TLOC color','Source IP','DST Public IP','DST Public Port','ENCAP','DETECT Multiplier','TX Interval','Site name','Batch info']
bfd_dict = {}

#Define method to Enhance Excel format
def ws_format(ws, tablename,width="auto"):
    max_row_num = ws.max_row
    #cell_list = []
    #for row in ws.iter_rows(min_row=2, max_row=max_row_num, min_col=1, values_only=False):
    #    for cell in row:
    #        cell_list.append(cell.coordinate)

    max_cell = ws.cell(row=ws.max_row, column=ws.max_column)
    max_cell_coordinat = max_cell.coordinate

    table_ref = "A1:"+ max_cell_coordinat
    tab = Table(displayName=tablename, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=True,)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    if width == 'auto':
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value*1.1
    else:
        i = 1
        while i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = width
            i += 1

#site id - site name mapping info
with open (site_list,'r') as f:
    site_id_list = f.read()
    site_id_list = site_id_list.split('\n')
    for line in site_id_list:
        temp_list = line.split(' ')
        site_id_dict[temp_list[0]] = temp_list[1]

#site name - batch list mapping info
with open(batch_list, 'r') as f:
    site_batch_list = f.read()
    site_batch_list = site_batch_list.split('\n')
    for line in site_batch_list:
        temp_list = line.split(' ')
        site_batch_dict[temp_list[0]] = temp_list[1]

#Go through all log files and put bfd sessions status & information into bfd_dict{'router name' = list(all_bfd_info)}
for file in files:
    file_path = os.path.join(path,file)
    with open(file_path) as f:
        raw_datas =f.read()
    raw_datas = raw_datas.split('\n')
    up_counts = 0
    down_counts = 0
    bfd_list = []
    for row in raw_datas:
        if row.startswith('10.'):
            row_list = row.split()
            if row_list[2] == 'up':
                up_counts += 1
            elif row_list[2] == 'down':
                down_counts +=1
            bfd_list.append(row_list)
    print(f'{file.split(".")[0]} have {len(bfd_list)} bfd sessions, {up_counts} is up, {down_counts} is down')
    mark_lines = [raw_datas.index(line) for line in raw_datas if '------' in line ]
    bfd_raw_datas = raw_datas[mark_lines[-1]+1:]
    i = 0
    bfd_sessions = []
    while i < len(bfd_raw_datas):
        if bfd_raw_datas[i].startswith('10'):
            bfd_temp_data_1 = [i for i in bfd_raw_datas[i].split(' ') if i != '']      #bfd session divided into 2 lines
            bfd_temp_data_2 = [i for i in bfd_raw_datas[i+1].split(' ') if i != '']
            if bfd_temp_data_1[-1] == '10': #some lines end with '10', the first part for TX interval value
                bfd_temp_data = bfd_temp_data_1[0:-1]
                bfd_temp_data.append(bfd_temp_data_1[-1]+bfd_temp_data_2[0])
                bfd_temp_data.extend(bfd_temp_data_2[1:])
                bfd_sessions.append(bfd_temp_data)
            elif bfd_temp_data_1[-1] == '1000': #some lines end with '1000', the full number of TX interval value
                bfd_temp_data = bfd_temp_data_1[:]
                bfd_temp_data.extend(bfd_temp_data_2[:])
                bfd_sessions.append(bfd_temp_data)
        i += 2
    bfd_dict[file.split('.')[0]] = bfd_sessions #put bfd session info into bfd_dict with router name as key, "CN02 R1 - before":"bfd_sessions"
    
#insert remvoed bfd session info into excel sheet
wb = Workbook()
del wb['Sheet']

i = 1
for router in router_list:
    router_before_bfd = bfd_dict[router + '- before'] 
    bfd_before_str = [ '&'.join(line[:-2]) for line in router_before_bfd] #uptime is different each time we show bfd sessions info, combine rest of info to str for set() comparison purpose
    router_after_bfd = bfd_dict[router + '- after']
    bfd_after_str = [ '&'.join(line[:-2]) for line in router_after_bfd]
    before_after_diff = list(set(bfd_before_str).symmetric_difference(set(bfd_after_str)))
    ws = wb.create_sheet(router)
    ws.append(headlines)
    for line in before_after_diff:
        bfd_session_info = line.split('&')
        bfd_session_info.append(site_id_dict[bfd_session_info[1]])
        if bfd_session_info[-1] in site_batch_dict.keys():
            bfd_session_info.append(site_batch_dict[bfd_session_info[-1]])
        ws.append(bfd_session_info)
    tablename = "table"+str(i)
    ws_format(ws,tablename)
    i += 1
wb.save(os.path.join(path_excel,'28th-June Removed BFD sessions on SDWAN routers check.xlsx'))
