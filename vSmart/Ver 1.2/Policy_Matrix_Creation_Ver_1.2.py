import os
from wsgiref import headers
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import BORDER_DOUBLE
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv

#get policy data from original vSmart file
def get_policy(keyword_1, keyword_2, textline):
    policy_content = []
    for row in textline:
        if keyword_1 in row:
            policy_content_start = textline.index(row)
            break
    for row in textline:
        if keyword_2 in row:
            policy_content_end = textline.index(row)
            break
    for row in textline[policy_content_start:policy_content_end]:
        policy_content.append(row)
    return policy_content

#get list data from original vSmart file
def get_list(keyword_1, keyword_2, textline):
    list_content = []
    for row in textline:
        if row == " lists\n":
            list_anchor = textline.index(row)
            break
    for row in textline[list_anchor:]:
        if keyword_1 in row:
            list_content_start = textline.index(row)
            break
    if keyword_2 != 'end':
        for row in textline[list_anchor:]:
            if keyword_2 in row:
                list_content_end = textline.index(row)
                break
    else:
        list_content_end = None
    for row in textline[list_content_start:list_content_end]:
        list_content.append(row)
    return list_content

#Define function to creat sheet and high light key word line
def sheet_create(list,excel_wb,sheet_name,highlight=None):
    fill_policy = PatternFill('solid', fgColor='FFFF00')
    fill_seq = PatternFill('solid', fgColor='33CCFF')
    ws = excel_wb.create_sheet(sheet_name)
    for row in list:
        row_list = row.split( )
        if row_list[0] == '!':
            continue
        ws.append(row_list)
    for cell in ws['A']:
        if cell.value == highlight:
            for cells in ws[cell.row]:
                cells.fill = fill_policy
        elif cell.value == 'sequence':
            cell.fill = fill_seq

#Define method to Enhance Excel format
def ws_format(ws, tablename, keyword = None, width='auto'):
    align = Alignment(horizontal='left', vertical='center')
    #side_thin = Side('thin')
    #side_double = Side('double')
    #side_dotted = Side('dotted')
    #content_border = Border(bottom=side_dotted, right=side_thin,left=side_thin)
    #header_border = Border(bottom=side_double, right=side_thin, left=side_thin)
    #bottom_border = Border(bottom=side_thin, right=side_thin,left=side_thin)
    fill = PatternFill('solid', fgColor='33CCFF')
    fill_seq = PatternFill('solid', fgColor='FFFF00')

    #Insert table with TableStyle "TableStyleLight9 - Blue", default with banded column and striped rows
    max_row_num = ws.max_row
    max_cell = ws.cell(row=ws.max_row, column=ws.max_column)
    max_cell_coordinat = max_cell.coordinate
    table_ref = "A1:"+ max_cell_coordinat
    tab = Table(displayName=tablename, ref=table_ref)
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=True,)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    #Set headline alignment to "Center"
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, values_only=False):
        for cell in row:
            cell.alignment = align
    #Highlight body lines including value "Sequence" & Predefined "KeyWord" 
    for row in ws.iter_rows(min_row=2, max_row=max_row_num, min_col=1, values_only=False):
        for cell in row:
            if cell.value != None:
                if 'sequence' in cell.value:
                    cell.fill = fill_seq
                if str(keyword) in cell.value:
                    cell.fill = fill
    
    #Adopt cell width to value length
    if width == 'auto':
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value*1.03
    else:
        i = 1
        while i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = width
            i += 1

#Define fuction to inset dict into excel
def insert_sheet(dict,wb,sheet_name, header_list=None):
    workbook = wb
    ws = workbook.create_sheet(sheet_name)
    if header_list != None:
        ws.append(header_list)
    for key in dict:
        temp_list = []     
        temp_list.append(key)
        for element in dict[key]:
            temp_list.append(element)
        ws.append(temp_list)
    return ws

#Insert list by column
def Insert_content_column(wb,policy_list,sheet_name, keyword = None, tag='no'):
    workbook = wb
    ws = workbook.create_sheet(sheet_name)
    column_num = 1
    keep_num = 1
    row_num = 1
    for list in policy_list:
        #print(len(list))
        #print('column_num is {}'.format(column_num))
        if tag == 'yes':
            if row_num <= 40:
                row_num = keep_num
            else:
                row_num = 1
        else:
            row_num =1
        for element in list:
            ws.cell(row=row_num, column=column_num).value = element
            row_num += 1
        if tag == 'yes':
            if row_num <= 40:
                keep_num = row_num
            else:
                column_num += 1
        else:
            column_num +=1
    ws_format(ws,sheet_name, keyword= keyword)

#Define file path
path = 'E:/Python/L-Oreal/vSmart/Ver 1.2/'
vsmart_config = 'vSmart.txt'
sitelist_file = 'site_list.txt'
file_path = path + vsmart_config
apply_policy_path = path + 'Policy_Matrix.xlsx'

sitelist_path = path + sitelist_file

with open(file_path, 'r', encoding = 'utf-8') as f:
    textline = f.readlines()
with open(sitelist_path, 'r', encoding = 'utf-8') as f:
    sitelist_list = f.readlines()

#define site_name & site_ID mapping diction site_id_mapping_dict
#get sitelist diction with {"site name" : "site ID"}
site_id_mapping_dict = {}

for row in sitelist_list:
    list = row.split( )
    site_id_mapping_dict[list[1]] = list[0]

#get site_list & apply_policy from original vsmart configuration
slaclass = get_policy('sla-class','data-policy',textline)
datapolicy = get_policy('data-policy', 'app-route-policy', textline)
aarpolicy= get_policy('app-route-policy', 'cflowd-template', textline)
cflowd = get_policy('cflowd-template', 'lists', textline)
data_prefix_list = get_list('data-prefix-list', 'tloc-list', textline)
tloc_list = get_list('tloc-list','app-list',textline)
app_list = get_list('app-list','color-list',textline)
color_list = get_list('color-list','site-list',textline)
site_list = get_list('site-list','  prefix-list ',textline)
prefix_list = get_list('  prefix-list ', 'control-policy', textline)
control_policy = get_list('control-policy', 'apply-policy', textline)
apply_policy = get_list('apply-policy', 'end', textline)

#Get site_list_name & site_id mapping information into diction sl_siteid_range_dict like {"site_list_name" : "site ID list"}
#for single site id, save as int type in the site ID list
#for site id range, save as list type with first number and last number
sl_siteid_range_dict = {}

index_list = [site_list.index(line) for line in site_list if 'site-list' in line]
#for site_list_line_1 in site_list:
#    if 'site-list' in site_list_line_1:
#       index_list.append(site_list.index(site_list_line_1))

i = 0
while i < len(index_list):
    siteid = []
    if i == len(index_list)-1:
        temp_sitelist = site_list[index_list[i]+1:-1]
    else:
        temp_sitelist = site_list[index_list[i]+1:index_list[i+1]-1]
    for id in temp_sitelist:
        temp_siteid = id.strip('!').split( ).pop()
        if temp_siteid == '':
            continue
        elif '-' in temp_siteid:
            siteid.append(temp_siteid.split('-'))
        else:
            siteid.append(temp_siteid)
    sl_siteid_range_dict[site_list[index_list[i]].split( ).pop()] = siteid
    i += 1

#for i in sl_siteid_range_dict:
#    print("{} is {} ".format(i,sl_siteid_range_dict[i]))
#Get site_list_name & site_policy mapping dict
#{site_list_name: [policy_name_1, policy_name2, ...]}
site_policies_dict = {}
sl_index = [apply_policy.index(i) for i in apply_policy if 'site-list' in i] #Get site-list index to locate each site list lines
i = 0
while i < len(sl_index):
    policy_list = ['','','','','']
    if i == len(sl_index)-1: #the last site list
        temp_policy_list = apply_policy[sl_index[i]:]
    else: #get policy attached to site list
        temp_policy_list = apply_policy[sl_index[i]:sl_index[i+1]]
    for site_policy in temp_policy_list:
        temp_data = site_policy.split( )
        if 'control-policy' in site_policy:
            if temp_data[2] == 'in':
                policy_list[0] = temp_data[1]
            else:
                policy_list[1] = temp_data[1]
        elif 'data-policy' in site_policy:
            policy_list[2] = temp_data[1]
            policy_list[3] = temp_data[2]
        elif 'app-route-policy' in site_policy:
            policy_list[4] = temp_data[1]
        elif 'site-list' in site_policy:
            site_list_name = site_policy.split( )
    site_policies_dict[site_list_name[1]] = policy_list
    i += 1
#for site_list_name in sl_siteid_range_dict:
#    print('{} is {}'.format(site_list_name,sl_siteid_range_dict[site_list_name]))
#Get site and site_list_name mapping information and put into Diction site_name_site_list_mapping_dict
site_name_site_list_mapping_dict = {}
none_policy_site_list = []

for site_name in site_id_mapping_dict:
    list_2 = []
    for site_list_name in sl_siteid_range_dict:
        if site_policies_dict.get(site_list_name) == None: #remove sites without attached policy
            if none_policy_site_list.count(site_list_name) < 2:
                none_policy_site_list.append(site_list_name) #Pick those sites without policy attached
            continue
        site_list_name_list = sl_siteid_range_dict[site_list_name]
        #print(site_list_name)
        #print(site_list_name_list)
        #print('site_list_name_list is {}'.format(site_list_name_list))
        for site_num in site_list_name_list:
            if isinstance(site_num, str):
                if int(site_id_mapping_dict[site_name]) == int(site_num):
                    list_2.append(site_list_name)
                    break
            else:
                if int(site_num[0]) <= int(site_id_mapping_dict[site_name]) <= int(site_num[1]):
                    list_2.append(site_list_name)
                    break
    site_name_site_list_mapping_dict[site_name] = list_2





#Get site_list_name_list
site_list_name_list_header = ['Site Name',]
for key in site_policies_dict:
    site_list_name_list_header.append(key)

#create a blank match list
i = 0
match_list = [None,]
while i <= len(site_list_name_list_header):
    match_list.append(None)
    i +=1
#Merge match list and site_name_list into a new dict
site_name_site_list_excel_dict = {}
for key_site_name in site_name_site_list_mapping_dict:
    match_list_copy = match_list[:]
    for element_site_list_1 in site_name_site_list_mapping_dict[key_site_name]:
        for element_site_list_2 in site_list_name_list_header:
            if element_site_list_1 == element_site_list_2:
                element_num = site_list_name_list_header.index(element_site_list_2)
                match_list_copy[element_num-1] = 'Yes'
    site_name_site_list_excel_dict[key_site_name] = match_list_copy

#Get site name and detail policy mapping information
sn_policies_mapping = {}
for key in site_name_site_list_mapping_dict:
    temp_list = [None,None,None,None,None]
    for value_1 in site_name_site_list_mapping_dict[key]:
        for value_2 in site_policies_dict[value_1]:
            if value_2 != '':
                if 'cp_FROM' in value_2:
                    temp_list[0] = value_2
                elif 'cp_TO' in value_2:
                    temp_list[1] = value_2
                elif 'all' in value_2 or 'from-tunnel' in value_2 or 'from-service' in value_2:
                    temp_list[3] = value_2
                elif '_arp_' in value_2:
                    temp_list[4] = value_2
                else:
                    temp_list[2] = value_2
    sn_policies_mapping[key] = temp_list

#Get content between 2 key words and return list
def capture_content_list(keyword, list):
    content_list = []
    for line_1 in list:
        if keyword in line_1:
            start_line = list.index(line_1)
            end_line = None
            for line_2 in list[start_line+1:]:               
                if keyword in line_2:
                    end_line = list.index(line_2)
                    break
            if end_line == None:
                end_line = len(list)
            temp_list = []
            for line_3 in list[start_line:end_line]:
                temp_list.append(line_3)
            content_list.append(temp_list)
    return content_list



wb = Workbook()
del wb['Sheet']


try:
    ws_site_list = insert_sheet(site_name_site_list_excel_dict,wb,'Sl_Matrix', site_list_name_list_header)
except:
    None
ws_format(ws_site_list,'SL_Matrix', 15)

policy_matrix_header = ['Site-list-Name', 'Control-policy In', 'Control-policy Out', 'Data-Policy', 'Type', 'AAR-policy']
ws_policy_matrix = insert_sheet(site_policies_dict,wb,'Policy-Matrix', policy_matrix_header)
ws_format(ws_policy_matrix, 'Policy-Matrix')

site_name_list_header = ['Site Name', 'Policy 1', 'Policy 2', 'Policy 3']
ws_site_name_list = insert_sheet(site_name_site_list_mapping_dict,wb,'site-name', site_name_list_header)
ws_format(ws_site_name_list, 'site-name')

sn_policy_header = ['Site Name', 'CP From', 'CP To', 'Data-Policy', 'DP Direction','AAR-Policy']
ws_site_name_policy = insert_sheet(sn_policies_mapping,wb,'SN list', sn_policy_header)
ws_format(ws_site_name_policy, "SN-list")

aarpolicy_list = capture_content_list('app-route-policy',aarpolicy)
Insert_content_column(wb,aarpolicy_list, 'app-route-policy', 'AAR-policy')
datapolicy_list = capture_content_list('data-policy', datapolicy)
Insert_content_column(wb,datapolicy_list,'data-policy','DP')
control_policy_list = capture_content_list('control-policy',control_policy)
Insert_content_column(wb,control_policy_list,'CP')
apply_policy_list = capture_content_list('site-list',apply_policy)
Insert_content_column(wb,apply_policy_list,'Apply-policy','site-list','yes')
data_prefix_list_list = capture_content_list('data-prefix-list',data_prefix_list)
Insert_content_column(wb,data_prefix_list_list,'Data-Prefix-list','data-prefix-list', 'yes')
prefix_list_list = capture_content_list('prefix-list',prefix_list,)
Insert_content_column(wb,prefix_list_list,'Prefix-list','prefix-list','yes')
sla_class_list = capture_content_list('sla-class',slaclass)
Insert_content_column(wb,sla_class_list,'Sla-class')

sheet_create(cflowd, wb,'Cflowd-Template')
tloc_list_list = capture_content_list('tloc-list',tloc_list)
Insert_content_column(wb, tloc_list_list,'Tloc-list','tloc-list','yes')
app_list_list = capture_content_list('app-list',app_list)
Insert_content_column(wb, app_list_list,'APP-list','app-list','yes')
color_list_list = capture_content_list('color-list', color_list)
Insert_content_column(wb,color_list_list,'Color-list','color-list','yes')
site_list_list = capture_content_list('site-list',site_list)
Insert_content_column(wb,site_list_list,'Site-list','site-list','yes')

wb.save(apply_policy_path)






