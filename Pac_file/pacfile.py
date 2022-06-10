import os
from typing import List
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import BORDER_DOUBLE
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.workbook import workbook

def ws_format(ws, keyword = None, width='auto'):
    align = Alignment(horizontal='left', vertical='center')
    side_thin = Side('thin')
    side_double = Side('double')
    side_dotted = Side('dotted')
    content_border = Border(bottom=side_dotted, right=side_thin,left=side_thin)
 #   header_border = Border(bottom=side_double, right=side_thin, left=side_thin)
    bottom_border = Border(bottom=side_thin, right=side_thin,left=side_thin)
    fill = PatternFill('solid', fgColor='33CCFF')
    fill_seq = PatternFill('solid', fgColor='FFFF00')

 #   for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, values_only=False):
 #       for cell in row:
 #           cell.fill = fill
 #           cell.alignment = align
 #           cell.border = header_border

    max_row_num = ws.max_row

    for row in ws.iter_rows(min_row=1, max_row=max_row_num-1, min_col=1, values_only=False):
        for cell in row:
            cell.alignment = align
            cell.border = content_border
            if cell.value != None:
                if 'return' in cell.value:
                    cell.fill = fill_seq
                if str(keyword) in cell.value:
                    cell.fill = fill

    for cell in ws[max_row_num]:
        cell.alignment = align
        cell.border = bottom_border
        if cell.value != None:
            if 'return' in cell.value:
                cell.fill = fill_seq
            elif str(keyword) in cell.value:
                cell.fill = fill
            elif 'default proxy' in cell.value:
                cell.fill = fill
    
    if width == 'auto':
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value*1.03
    else:
        i = 1
        while i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = width
            i += 1

#Define fuction to inset dict into excel
def insert_sheet(dict,wb,sheet_name, header_list=None):
    workbook = wb
    ws = workbook.create_sheet(sheet_name)
    if header_list != None:
        ws.append(header_list)
    for key in dict:
        temp_list = []     
        temp_list.append(key)
        for element in dict[key]:
            temp_list.append(element)
        ws.append(temp_list)
    return ws
#Insert list by column
def Insert_content_column(wb,policy_list,sheet_name, keyword = None, tag='no'):
    workbook = wb
    ws = workbook.create_sheet(sheet_name)
    column_num = 1
    keep_num = 1
    row_num = 1
    for list in policy_list:
        #print(len(list))
        #print('column_num is {}'.format(column_num))
        if isinstance(list, List):
            if tag == 'yes':
                if row_num <= 40:
                    row_num = keep_num
                else:
                    row_num = 1
            else:
                row_num =1
            for element in list:
                ws.cell(row=row_num, column=column_num).value = element
                row_num += 1
            if tag == 'yes':
                if row_num <= 40:
                    keep_num = row_num
                else:
                    column_num += 1
            else:
                column_num +=1
        else:
            ws.cell(row=row_num, column=column_num).value = list
            row_num += 1
            keep_num = row_num
    ws_format(ws,keyword)

def get_pac_list(pac_file):
    #Split Pac_file by if(),{}
    final_list = []
    while len(pac_file) > 3:
        for line_1 in pac_file:
            if '{' in line_1:
                start_line = pac_file.index(line_1)
                tag = None
                #print('start_line is {}'.format(start_line))
                while tag != 'end':
                    for line_2 in pac_file[start_line+1:]:
                        if '{' in line_2:
                            tag = 'continue'
                            start_line = pac_file.index(line_2)
                            #print('tag is {}, start_line is {}'.format(tag,start_line))
                        if '}' in line_2:
                            tag = 'end'
                            end_line = pac_file.index(line_2)
                            #print('tag is {}'.format(tag))
                            break
    #Get intial start_line and end_line
                temp_list = []
                #print('start_line is {} end_line is {}'.format(start_line,end_line))
                real_start_line = 0
                for line_4 in pac_file[:start_line]:
                    if 'if' in line_4:
                        real_start_line = pac_file.index(line_4)
                #print('real_start_line is {}'.format(real_start_line))
            #if no 'if' beofer {}, consider as intial {}, get all scripts
                if real_start_line == 0:
                    for line_3 in pac_file[start_line:end_line+2]:
                        temp_list.append(line_3)
                        del pac_file[pac_file.index(line_3)]
                    final_list.append(temp_list)
                else:
            #If have "if" before {}, get all scripts with if () {} return
                    #print('real_start_line is {}, end_line is {}'.format(real_start_line,end_line+1))
                    for line_3 in pac_file[real_start_line:end_line+1]:
                        temp_list.append(line_3)
                        del pac_file[pac_file.index(line_3, real_start_line, end_line+1)]
                    final_list.append(temp_list)
                #print('len(pac_file) is {}'.format(len(pac_file)))
    return final_list

#search desigated IP address over full list without {} parts
def get_proxy_list(ip_address, pac_list):
    tag = None
    proxy_list = []
    num = 0
    end_line = 0
    for line_1 in pac_list:
        if ip_address in line_1 and 'return' in line_1: 
            print('end_line is {}, num is {} length is {}'.format(end_line, num, len(pac_list)))
            end_line = pac_list.index(line_1, num, len(pac_list))
            num = end_line+1
            start_line = 0
            print('end_line is {}, num is {} length is {}'.format(end_line, num, len(pac_list)))
            for line_2 in pac_list[:end_line]:
                if 'if (' in line_2:
                    start_line = pac_list.index(line_2,start_line+1,end_line)
            temp_list = []
            print('start_line is {}, end_line is {}'.format(start_line,end_line))
            for line_3 in pac_list[start_line:end_line]:
                if 'return' in line_3:
                    tag = 'final'
                    break
                else:
                    temp_list.append(line_3)
            temp_list.append(pac_list[end_line])

            if tag == 'final':
                proxy_list.append('Proxy ' + ip_address + ' is default proxy')
            else:
                proxy_list.append(temp_list)
    return proxy_list

#serrch desigated IP addres over {} part
def get_proxy_list_2(ip_address, pac_list):
    for line in pac_list:
        if ip_address in line:
            return pac_list

#Create a new excel workbook
wb = Workbook()
del wb['Sheet']

#define path for pac files
path = 'I:/Python/Pac_file/Pac list/'
files = os.listdir(path)
excel = "I:/Python/Pac_file/pac_file_check.xlsx"
#excel_path = path + excel

#Get IP addres 
ip_address = str(input('please input the proxy IP addres you want to check in PAC files \n'))

#Going thourgh all PAC files in the folder
for file in files:
    file_path = path + file

    with open(file_path, 'r', encoding='utf=8') as f:
        pac_file = f.readlines()
        print('Now is checking {}'.format(file_path))

    pac_list = get_pac_list(pac_file)


#Check full pac list without {} part
    proxy_list = get_proxy_list(ip_address, pac_list[len(pac_list)-1])

#Check {} part
    if len(pac_list) > 1:
        for list in pac_list[:len(pac_list)-1]:
            temp_list = get_proxy_list_2(ip_address, list)
            if temp_list != None:
                proxy_list.append(temp_list)
    if len(proxy_list) < 1:
        print('No Proxy IP address {} in PAC file {}'.format(ip_address,file))
    else:
        ws = Insert_content_column(wb,proxy_list,file,'if (','yes')

print('Search completed, generating excel file')
wb.save(path+'PAC_file_check_'+ip_address+'.xlsx')

        

            