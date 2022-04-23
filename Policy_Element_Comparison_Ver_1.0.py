import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import BORDER_DOUBLE
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

#get policy data from original vSmart file
def get_policy(keyword_1, keyword_2, textline):
    policy_content = []
    for row in textline:
        if keyword_1 in row:
            policy_content_start = textline.index(row)
            break
    for row in textline:
        if keyword_2 in row:
            policy_content_end = textline.index(row)
            break
    for row in textline[policy_content_start:policy_content_end]:
        policy_content.append(row)
    return policy_content

#get list data from original vSmart file
def get_list(keyword_1, keyword_2, textline):
    list_content = []
    for row in textline:
        if row == " lists\n":
            list_anchor = textline.index(row)
            break
    for row in textline[list_anchor:]:
        if keyword_1 in row:
            list_content_start = textline.index(row)
            break
    if keyword_2 != 'end':
        for row in textline[list_anchor:]:
            if keyword_2 in row:
                list_content_end = textline.index(row)
                break
    else:
        list_content_end = None
    for row in textline[list_content_start:list_content_end]:
        list_content.append(row)
    return list_content

#Define function to creat sheet and high light key word line
def sheet_create(list,excel_wb,sheet_name,highlight=None):
    fill_policy = PatternFill('solid', fgColor='FFFF00')
    fill_seq = PatternFill('solid', fgColor='33CCFF')
    ws = excel_wb.create_sheet(sheet_name)
    for row in list:
        row_list = row.split( )
        if row_list[0] == '!':
            continue
        ws.append(row_list)
    for cell in ws['A']:
        if cell.value == highlight:
            for cells in ws[cell.row]:
                cells.fill = fill_policy
        elif cell.value == 'sequence':
            cell.fill = fill_seq

#Define function to Enhance Excel format
def ws_format(ws, width='auto'):
    align = Alignment(horizontal='left', vertical='center')
    side_thin = Side('thin')
    side_double = Side('double')
    side_dotted = Side('dotted')
    content_border = Border(bottom=side_dotted, right=side_thin,left=side_thin)
    header_border = Border(bottom=side_double, right=side_thin, left=side_thin)
    bottom_border = Border(bottom=side_thin, right=side_thin,left=side_thin)
    fill = PatternFill('solid', fgColor='33CCFF')

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, values_only=False):
        for cell in row:
            cell.fill = fill
            cell.alignment = align
            cell.border = header_border

    max_row_num = ws.max_row

    for row in ws.iter_rows(min_row=2, max_row=max_row_num-1, min_col=1, values_only=False):
        for cell in row:
            cell.alignment = align
            cell.border = content_border

    for cell in ws[max_row_num]:
        cell.alignment = align
        cell.border = bottom_border
    
    if width == 'auto':
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value*1.1
    else:
        i = 1
        while i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = width
            i += 1

#Define fuction to inset dict into excel
def insert_sheet(dict,wb,sheet_name, header_list=None):
    workbook = wb
    ws = workbook.create_sheet(sheet_name)
    if header_list != None:
        ws.append(header_list)
    for key in dict:
        temp_list = []     
        temp_list.append(key)
        for element in dict[key]:
            temp_list.append(element)
        ws.append(temp_list)
    return ws

#Define file path
path = './'
file_path = path + 'vSmart.txt'
sitelist_path = path + 'site_list.txt'

with open(file_path, 'r', encoding = 'utf-8') as f:
    textline = f.readlines()
with open(sitelist_path, 'r', encoding = 'utf-8') as f:
    sitelist_list = f.readlines()

#define site_name & site_ID mapping diction site_id_mapping_dict
#get sitelist diction with {"site name" : "site ID"}
site_id_mapping_dict = {}

for row in sitelist_list:
    list = row.split( )
    site_id_mapping_dict[list[1]] = list[0]

#get site_list & apply_policy from original vsmart configuration
slaclass = get_policy('sla-class','data-policy',textline)
datapolicy = get_policy('data-policy', 'app-route-policy', textline)
aarpolicy= get_policy('app-route-policy', 'cflowd-template', textline)
cflowd = get_policy('cflowd-template', 'lists', textline)
data_prefix_list = get_list('data-prefix-list', 'tloc-list', textline)
tloc_list = get_list('tloc-list','app-list',textline)
app_list = get_list('app-list','color-list',textline)
color_list = get_list('color-list','site-list',textline)
site_list = get_list('site-list','  prefix-list ',textline)
prefix_list = get_list('  prefix-list ', 'control-policy', textline)
control_policy = get_list('control-policy', 'apply-policy', textline)
apply_policy = get_list('apply-policy', 'end', textline)

#Get site_list_name & site_id mapping information into diction sl_siteid_range_dict like {"site_list_name" : "site ID list"}
#for single site id, save as int type in the site ID list
#for site id range, save as list type with first number and last number
sl_siteid_range_dict = {}
for site_list_line in site_list:
    if 'site-list' in site_list_line:
        #define start_line & end_line to locate certain site_list range
        siteid = []
        site_list_name = site_list_line.split( )
        start_line = site_list.index(site_list_line)
        if start_line < len(site_list)-3:
            for element_2 in site_list[start_line+1:]:
                if 'site-list' in element_2:
                    end_line = site_list.index(element_2)
                    break
        else:
            end_line = len(site_list)
#        print('{}, start line is {}, end line is {}'.format(site_list_name[1],start_line,end_line))
        for element_3 in site_list[start_line:end_line]:
            if 'site-id' in element_3:
                temp_data = element_3.split( )
                if '-' in temp_data[1]:
                    siteid.append(temp_data[1].split('-'))
                else: 
                    siteid.append(temp_data[1])
        sl_siteid_range_dict[site_list_name[1]] = siteid
#Get site_list_name & site_policy mapping dict
#{site_list_name: [policy_name_1, policy_name2, ...]}
site_policies_dict = {}
for line in apply_policy:
    policy_list = ['','','','','']
    if 'site-list' in line:
        #define start_line & end_line to locate certain site_list range
        site_list_name = line.split( )
        start_line = apply_policy.index(line)
        if start_line < len(apply_policy)-10:
            for line_2 in apply_policy[start_line+1:]:
                if 'site-list' in line_2:
                    end_line = apply_policy.index(line_2)
                    break
        else:
            end_line = len(apply_policy)
#        print('{}, start line is {}, end line is {}'.format(site_list_name[1],start_line,end_line))
        for site_policy in apply_policy[start_line:end_line]:
            temp_data = site_policy.split( )
            if 'control-policy' in site_policy:
                if temp_data[2] == 'in':
                    policy_list[0] = temp_data[1]
                else:
                    policy_list[1] = temp_data[1]
            elif 'data-policy' in site_policy:
                policy_list[2] = temp_data[1]
                policy_list[3] = temp_data[2]
            elif 'app-route-policy' in site_policy:
                policy_list[4] = temp_data[1]
        site_policies_dict[site_list_name[1]] = policy_list

#for site_list_name in sl_siteid_range_dict:
#    print('{} is {}'.format(site_list_name,sl_siteid_range_dict[site_list_name]))
#Get site and site_list_name mapping information and put into Diction site_name_site_list_mapping_dict
site_name_site_list_mapping_dict = {}
none_policy_site_list = []
for site_name in site_id_mapping_dict:
    list_2 = []
    for site_list_name in sl_siteid_range_dict:
        if site_policies_dict.get(site_list_name) == None: #remove sites without attached policy
            if none_policy_site_list.count(site_list_name) < 2:
                none_policy_site_list.append(site_list_name) #Pick those sites without policy attached
            continue
        site_list_name_list = sl_siteid_range_dict[site_list_name]
        for site_num in site_list_name_list:
            if isinstance(site_num, str):
                if int(site_id_mapping_dict[site_name]) == int(site_num):
                    list_2.append(site_list_name)
                    break
            else:
                if int(site_num[0]) <= int(site_id_mapping_dict[site_name]) <= int(site_num[1]):
                    list_2.append(site_list_name)
                    break
    site_name_site_list_mapping_dict[site_name] = list_2


#Get site_list_name_list
site_list_name_list_header = ['Site Name',]
for key in site_policies_dict:
    site_list_name_list_header.append(key)

#create a blank match list
i = 0
match_list = [None,]
while i <= len(site_list_name_list_header):
    match_list.append(None)
    i +=1
#Merge match list and site_name_list into a new dict
site_name_site_list_excel_dict = {}
for key_site_name in site_name_site_list_mapping_dict:
    match_list_copy = match_list[:]
    for element_site_list_1 in site_name_site_list_mapping_dict[key_site_name]:
        for element_site_list_2 in site_list_name_list_header:
            if element_site_list_1 == element_site_list_2:
                element_num = site_list_name_list_header.index(element_site_list_2)
                match_list_copy[element_num-1] = 'Yes'
    site_name_site_list_excel_dict[key_site_name] = match_list_copy

#Get site name and detail policy mapping information
sn_policies_mapping = {}
for key in site_name_site_list_mapping_dict:
    temp_list = [None,None,None,None,None]
    for value_1 in site_name_site_list_mapping_dict[key]:
        for value_2 in site_policies_dict[value_1]:
            if value_2 != '':
                if 'cp_FROM' in value_2:
                    temp_list[0] = value_2
                elif 'cp_TO' in value_2:
                    temp_list[1] = value_2
                elif 'all' in value_2 or 'from-tunnel' in value_2 or 'from-service' in value_2:
                    temp_list[3] = value_2
                elif '_arp_' in value_2:
                    temp_list[4] = value_2
                else:
                    temp_list[2] = value_2
    sn_policies_mapping[key] = temp_list
#Insert Dicts into new Sheets

def capture_content_list(keyword_1, list, keyword_2='data-policy'):
    end_line = None
    temp_list = []
    for line_1 in list:
        if keyword_1 in line_1:
            start_line = list.index(line_1)
            for line_2 in list[start_line+1:]:
                if keyword_2 in line_2:
                    end_line = list.index(line_2)
                    break
            if end_line == None:
                end_line = len(list)
    for line_3 in list[start_line:end_line]:
        if '!' not in line_3:
            temp_list.append(line_3)
    return temp_list

sn_input = input('please input the site name you want to check:\n')

while sn_policies_mapping.get(sn_input) == None:
    sn_input = input('No site name found, please re-input the site name:\n')

sn_cp_from = capture_content_list(sn_policies_mapping[sn_input][0], control_policy, 'control-policy')
sn_cp_to = capture_content_list(sn_policies_mapping[sn_input][1], control_policy, 'control-policy')
sn_cp_dp = capture_content_list(sn_policies_mapping[sn_input][2], datapolicy)
print(sn_policies_mapping[sn_input][4])
sn_cp_aar = capture_content_list(sn_policies_mapping[sn_input][4], aarpolicy,'app-route-policy')

wb = Workbook()
ws = wb.active

def insert_policy_data(list,ws,column='A'):
    i = 1
    while i < len(list):
        ws[column+str(i)] = list[i-1]
        i += 1
insert_policy_data(sn_cp_from,ws,)
insert_policy_data(sn_cp_to,ws,'B')
insert_policy_data(sn_cp_dp,ws,'C')
insert_policy_data(sn_cp_aar,ws,'D')
ws_format(ws)

wb.save(path+sn_input+'.xlsx')

            






